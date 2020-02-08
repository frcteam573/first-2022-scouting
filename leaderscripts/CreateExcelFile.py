#Opens a template workbook and creates a copy of a template worksheet
#within a copy of the workbook

#Import libraries
import os
from openpyxl import load_workbook
import openpyxl
import shutil
from datetime import datetime
import json
import sqlite3
#from PIL import Image
#import PIL

#row number :
r = 24


#Print Script Start
print ('!!!!!!!!! Script Start !!!!!!!!!')

#Changes Working Directory To a Known Location
#os.chdir('C:\Users\es2433\Desktop\PV_PythonSandbox\ExcelBookLayout')

#Copy Excel file from template file with datetime stamp
CurrDateTime = str(datetime.now())
CurrDateTime = CurrDateTime.replace(':','-')
CurrDateTime = CurrDateTime.replace(' ','_')
CurrDateTime = CurrDateTime.replace('.','-')
shutil.copy('ScoutingData_template.xlsx','ScoutingData_'+CurrDateTime+'.xlsx')

#Connect To Match Database
db = sqlite3.connect('C:\\dev\\scouting_app_2019\\2019_Scouting_App\\db.sqlite3')
c = db.cursor()

#Connect To Pit Database
##db_pit = sqlite3.connect('CombinedPitData')
##c_pit = db_pit.cursor()

#Opens Excel Spreadsheet
wb = load_workbook('ScoutingData_'+CurrDateTime+'.xlsx')

#Find template worksheet by name
ws_template = wb['template']#.get_sheet_by_name(name = 'template')

# ----------------------------------------
#Random code isn't needed

##my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
##my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
##ws1["A1"].value = 'BLAH'
##ws1["A1"].fill = my_fill

#----------------------------------------

#Read in team numbers from team list
wsteam_list = wb['team_list']#.get_sheet_by_name(name = 'team_list')
TeamList=[]
TeamNameList=[]

#Populate Team List Based Upon Excel Data
for rowindex in range(wsteam_list.max_row):
    TeamList.append(str(wsteam_list.cell(row = rowindex + 1,column = 1).value))
    TeamNameList.append(str(wsteam_list.cell(row = rowindex + 1,column = 2).value))

#Loop through team list creating worksheet for each one
for inc, team in enumerate(TeamList):
#Create copy of template sheet
    ws2 = wb.copy_worksheet(ws_template)

#Rename Newly Copied Sheet
    ws2.title = team
#----------------------------------------

    #Fill out data on sheet
    ws2.cell(row=1,column=2).value = team
    ws2.cell(row=2,column=2).value = TeamNameList[inc]
    
    #Put in pit scouting data
    c.execute("SELECT * FROM pitscout_pitscout WHERE team_num="+team)
    rows_pit = c.fetchall()

    for inc3, row_pit in enumerate(rows_pit):
        ws2.cell(row=inc3+5,column=2).value = row_pit[3]
        ws2.cell(row=inc3+6,column=2).value = row_pit[7]
        ws2.cell(row=inc3+7,column=2).value = row_pit[6]
        ws2.cell(row=inc3+8,column=2).value = row_pit[14]
        ws2.cell(row=inc3+9,column=2).value = row_pit[12]
        ws2.cell(row=inc3+10,column=2).value = row_pit[8]
        ws2.cell(row=inc3+11,column=2).value = row_pit[9]
        ws2.cell(row=inc3+12,column=2).value = row_pit[10]
        ws2.cell(row=inc3+13,column=2).value = row_pit[11]

    #Put in match scouting data
    c.execute("SELECT * FROM matchscout_matchscout WHERE team_num="+team)
    rows = c.fetchall()
    for inc2, row in enumerate(rows):
        #SUBTRACT 1 FROM EACH
        #name 2
        #team_num 3
        #match_num 4
        #alliance 5
        #starting 6
        #s_hatches_1 7
        #s_hatches_2 8
        #s_hatches_3 9
        #s_cargo_1 10
        #s_cargo_2 11
        #s_cargo_3 12
        #t_hatches_1 13
        #t_hatches_2 14
        #t_hatches_3 20
        #t_cargo_1 15
        #t_cargo_2 16
        #t_cargo_3 17
        #ending 18
        #comments 19
        
        
        #somedata cleanup
        
        #Put Data into proper cells on worksheet
        ws2.cell(row=inc2+r,column=1).value = row[3]
        ws2.cell(row=inc2+r,column=2).value = row[4]
        ws2.cell(row=inc2+r,column=3).value = row[1]
        ws2.cell(row=inc2+r,column=4).value = row[5]
        ws2.cell(row=inc2+r,column=5).value = row[6]
        ws2.cell(row=inc2+r,column=6).value = row[7]
        ws2.cell(row=inc2+r,column=7).value = row[8]
        ws2.cell(row=inc2+r,column=8).value = row[9]
        ws2.cell(row=inc2+r,column=9).value = row[10]
        ws2.cell(row=inc2+r,column=10).value = row[11]
        ws2.cell(row=inc2+r,column=11).value = row[12]
        ws2.cell(row=inc2+r,column=12).value = row[13]
        ws2.cell(row=inc2+r,column=13).value = row[19]
        ws2.cell(row=inc2+r,column=14).value = row[14]
        ws2.cell(row=inc2+r,column=15).value = row[15]
        ws2.cell(row=inc2+r,column=16).value = row[16]
        ws2.cell(row=inc2+r,column=17).value = row[17]
        #ws2.cell(row=inc2+r,column=18).value = row[3]

###Put in picture ----------------------------------
##    if (os.path.isfile('Pic/'+team+'.jpg')): #Check to make sure picture exists
##        teamnum = team
##    else:
##        teamnum ='default' #Stand in Photo For Teams With Missing Picture
##
##    #print teamnum
##    #anchor = 'G1'
##    
##    im = Image.open('Pic/'+teamnum+'.jpg')
##    image2 = im.resize((250,250),Image.ANTIALIAS)
##    image2.save(teamnum + "_thumbnail.jpg", "JPEG")
##    img = openpyxl.drawing.image.Image(teamnum + "_thumbnail.jpg")
##    ws2.add_image(img,'G3')
# ------------------------------------------------------

    # Add in Summary Page Info
    ws_sum = wb['Summary']#.get_sheet_by_name(name = 'Summary')

    #Adds information to summary sheet from team worksheet
    ws_sum.cell(row = inc+9,column = 1).value = "='"+team+"'!"+"B1" #Team Number

    #AVG
    ws_sum.cell(row = inc+9,column = 2).value = "='"+team+"'!"+"D14" #Autos
    ws_sum.cell(row = inc+9,column = 3).value = "='"+team+"'!"+"E14"
    ws_sum.cell(row = inc+9,column = 4).value = "='"+team+"'!"+"F14"
    ws_sum.cell(row = inc+9,column = 5).value = "='"+team+"'!"+"G14"

    ws_sum.cell(row = inc+9,column = 6).value = "='"+team+"'!"+"I14" #Tele
    ws_sum.cell(row = inc+9,column = 7).value = "='"+team+"'!"+"J14"
    ws_sum.cell(row = inc+9,column = 8).value = "='"+team+"'!"+"K14"
    ws_sum.cell(row = inc+9,column = 9).value = "='"+team+"'!"+"L14"

    ws_sum.cell(row = inc+9,column = 10).value = "='"+team+"'!"+"M14" #Endgame  
    ws_sum.cell(row = inc+9,column = 11).value = "='"+team+"'!"+"N14"
    ws_sum.cell(row = inc+9,column = 12).value = "='"+team+"'!"+"O14"
    ws_sum.cell(row = inc+9,column = 13).value = "='"+team+"'!"+"P14"
    ws_sum.cell(row = inc+9,column = 14).value = "='"+team+"'!"+"Q14"

    #MAX
    ws_sum.cell(row = inc+9,column = 15).value = "='"+team+"'!"+"D15" #Autos
    ws_sum.cell(row = inc+9,column = 16).value = "='"+team+"'!"+"E15"
    ws_sum.cell(row = inc+9,column = 17).value = "='"+team+"'!"+"F15"
    
    ws_sum.cell(row = inc+9,column = 18).value = "='"+team+"'!"+"I15" #Tele
    ws_sum.cell(row = inc+9,column = 19).value = "='"+team+"'!"+"J15"
    ws_sum.cell(row = inc+9,column = 20).value = "='"+team+"'!"+"K15"
    ws_sum.cell(row = inc+9,column = 21).value = "='"+team+"'!"+"L15"

    #MIN
    ws_sum.cell(row = inc+9,column = 22).value = "='"+team+"'!"+"D16" #Autos
    ws_sum.cell(row = inc+9,column = 23).value = "='"+team+"'!"+"E16"
    ws_sum.cell(row = inc+9,column = 24).value = "='"+team+"'!"+"F16"

    ws_sum.cell(row = inc+9,column = 25).value = "='"+team+"'!"+"I16" #Tele
    ws_sum.cell(row = inc+9,column = 26).value = "='"+team+"'!"+"J16"
    ws_sum.cell(row = inc+9,column = 27).value = "='"+team+"'!"+"K16"
    ws_sum.cell(row = inc+9,column = 28).value = "='"+team+"'!"+"L16"

    ws_sum.cell(row = inc+9,column = 29).value = "='"+team+"'!"+"B5" #Pit Scout
    ws_sum.cell(row = inc+9,column = 30).value = "='"+team+"'!"+"B6"
    #ws["A1"] = "=SUM('1'!B1 + '2'!B1 )" #Example
    
#Save workbook
wb.save('ScoutingData_'+CurrDateTime+'.xlsx')

#Print Script End
print ('!!!!!!!!! File Created !!!!!!!!!')
